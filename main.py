#to do: 
# error handling  
    # make sure no file opening errors such as trying to get non-existent data
    # All team summary only runs if excel isnt open
    # remove copy button

# double chekc 4 mistakes
# comment code oh god please
import pandas as pd
import tkinter as tk
from tkinter import filedialog as fd
from tkinter import ttk
import openpyxl

root = tk.Tk()
root.geometry('500x500')
root.title("Summarizer App")
root.resizable(False, False)
 
matchAmount, data, climbing = tk.StringVar(value="6"), "", ''
avgScore, avgAuto, avgTeleop, CL1A, CL2A, CL3A, CL4A, CPA, CNA, CL1, CL2, CL3, CL4, CP, CN, defence, climbAcc, aMissed, tMissed, totalCycles = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
deepClimb, shallowClimb,  reefAlgae = False, False, False

#function to read sheet and put data into a list
def selectFile():
    global data, file
    filetypes = (
        ('Excel files', '*.xlsx'),
    )
    #selecting file
    file = fd.askopenfilename(
        title='Open a file',
        filetypes=filetypes)
    file = file.replace("/","\\")
    fileText.config(text=file)
    getDatafromFile()
    errorOutput.configure(text="")


def getDatafromFile():
    
    global data, file
    try:
        data = (pd.read_excel(file,skiprows=0, sheet_name="Scout Data")).values.tolist()

        i = 0
        while i < len(data):
            j = i+1
            while j < len(data):
                if data[i] == data[j]:
                    del data[j]
                else:
                    j+=1
            i+=1
    except:
        errorOutput.configure(text= "You have not put in a file")



    #print(data)

def fullSummary():
    global file, avgAuto, avgTeleop, data, matchAmount, climbing, avgScore, CL1A, CL2A, CL3A, CL4A, CPA, CNA, CL1, CL2, CL3, CL4, CP, CN, defence, climbAcc, aMissed, tMissed, totalCycles, deepClimb, shallowClimb, reefAlgae, rawData
    tempList = data

    errorOutput.configure(text="")

    try: 
        workbook = openpyxl.load_workbook(file)
        sheet = workbook["FullSummary"]

        for row in sheet.iter_rows(min_row=3, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        i = 0
        while len(tempList) >= 1:
            team = tempList[0][2]
            summarizeData(team)
            #push data to spreadsheet
            for j in range(len(rawData)):
                sheet.cell(row=i+3, column=j+1,value=rawData[j])
            i+=1
            #print(rawData)

            j = 0
            while j < len(tempList):
                if team == tempList[j][2]:
                    del tempList[j]
                    j = 0
                else:
                    j +=1

        workbook.save(file)
    except NameError:
        errorOutput.configure(text= "Make sure you have the correct File")
    except openpyxl.utils.exceptions.InvalidFileException:
        errorOutput.configure(text= "Make sure you have the correct File")
    except PermissionError:
        errorOutput.configure(text="Please close the file and try again")

#getting data
def summarizeData(team):
    global data, matchAmount, climbing, avgScore, avgAuto, avgTeleop,  CL1A, CL2A, CL3A, CL4A, CPA, CNA, CL1, CL2, CL3, CL4, CP, CN, defence, climbAcc, aMissed, tMissed, totalCycles, deepClimb, shallowClimb, reefAlgae, rawData
    climbing = ''
    avgScore, CL1A, CL2A, CL3A, CL4A, CPA, CNA, CL1, CL2, CL3, CL4, CP, CN, defence, climbAcc, aMissed, tMissed, totalCycles = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    deepClimb, shallowClimb,  reefAlgae = False, False, False
    teamData =  []

    getDatafromFile()

    for i in range(len(data)):
        if data[i][2] == team:
            teamData.append(data[i])

    if (matchAmount.get() == "6" and len(teamData) > 6):
        teamData = teamData[len(teamData)-6:len(teamData)]

    #print(teamData)

    #organizing and outputing stats
    for row in range(len(teamData)):
        CL1A += (teamData[row][4])
        CL2A += (teamData[row][5])
        CL3A += (teamData[row][6])       
        CL4A += (teamData[row][7])
        CPA += (teamData[row][8])
        CNA += (teamData[row][9])
        aMissed += (teamData[row][10])
        
        avgAuto += (teamData[row][4] *3) + (teamData[row][5] *4) + (teamData[row][6] *6) + (teamData[row][7] *7) + (teamData[row][8] *6) + (teamData[row][9] *4)

        if (teamData[row][4] > 1 or teamData[row][5] > 1 or teamData[row][6] > 1 or teamData[row][7] or teamData[row][8] > 1 or teamData[row][9] > 1):
            avgAuto +=3

        CL1 += (teamData[row][11])           
        CL2 += (teamData[row][12])
        CL3 += (teamData[row][13])
        CL4 += (teamData[row][14])
        CP += (teamData[row][15])        
        CN += (teamData[row][16])
        tMissed += (teamData[row][17])

        avgTeleop += (teamData[row][11]*2) + (teamData[row][12]*3) + (teamData[row][13]*4) + (teamData[row][14]*5) + (teamData[row][15]*6) + (teamData[row][16]*4)

        totalCycles += (teamData[row][4]) + (teamData[row][5])+(teamData[row][6])+(teamData[row][7])+(teamData[row][8])+(teamData[row][9])+(teamData[row][11])+(teamData[row][12])+(teamData[row][13])+(teamData[row][14])+(teamData[row][15])+(teamData[row][16])
        
        if (teamData[row][18]=='P'):
            avgTeleop +=2
        elif (teamData[row][18]=='D'):
            avgTeleop += 12
            climbAcc += 1
            deepClimb = True
        elif (teamData[row][18]=='S'):
            avgTeleop += 6
            climbAcc += 1
            shallowClimb = True

        if (teamData[row][19]):
            defence += 1    

        if (teamData[row][20]):
            reefAlgae = True

    if (matchAmount.get() == "6" and len(teamData) >= 6):
        divisor = 6
    elif len(teamData)!=0:
        divisor = len(teamData)
    else: 
        divisor = 1
    
    avgScore = round((avgAuto+avgTeleop)/divisor,2)
    avgAuto = round(avgAuto/divisor,2)
    avgTeleop = round(avgTeleop/divisor,2)

    CL1A = round(CL1A/divisor,2)
    CL2A = round(CL2A/divisor,2)
    CL3A = round(CL3A/divisor,2)
    CL4A = round(CL4A/divisor,2)
    CPA = round(CPA/divisor,2)
    CNA = round(CNA/divisor,2)
    aMissed = (round(aMissed/divisor,2)*100)

    CL1 = round(CL1/divisor,2)
    CL2 = round(CL2/divisor,2)
    CL3 = round(CL3/divisor,2)
    CL4 = round(CL4/divisor,2)
    CN = round(CN/divisor,2)
    CP = round(CP/divisor,2)
    tMissed = (round(tMissed/divisor,2)*100)

    defence = (round(defence/divisor,2)*100)
    climbAcc = round(climbAcc/divisor, 2)

    if (deepClimb and shallowClimb):
        climbing = 'B'
    elif (deepClimb):
        climbing = 'D'
    elif (shallowClimb):
        climbing = 'S'
    else:
        climbing = 'N'
    if len(teamData) != 0:
        rawData = [team, CL1A, CL2A,  CL3A,  CL4A,  CPA,  CNA,  aMissed, avgAuto,  CL1,  CL2,  CL3,  CL4,  CP,  CN,  tMissed, avgTeleop, defence,  climbAcc,  climbing, reefAlgae, avgScore]
    else:
        if len(data) >1:
            errorOutput.config(text="a team does not have data")
        rawData = [team]

def getData():
    global  data, matchAmount, climbing, avgScore, avgAuto, avgTeleop, CL1A, CL2A, CL3A, CL4A, CPA, CNA, CL1, CL2, CL3, CL4, CP, CN, defence, climbAcc, aMissed, tMissed, totalCycles, deepClimb, shallowClimb, reefAlgae
    tableList = []

    errorOutput.configure(text="")
    tableWindow = tk.Toplevel()
    tableWindow.title("Summarized Data")
    tableWindow.config(width=1000, height=400)

    if submitTeamNum1.get().isdigit():
        summarizeData(int(submitTeamNum1.get()))
        tableList.append(rawData)
    if submitTeamNum2.get().isdigit():
        summarizeData(int(submitTeamNum2.get()))
        tableList.append(rawData)
    if submitTeamNum3.get().isdigit():
        summarizeData(int(submitTeamNum3.get()))
        tableList.append(rawData)
    if submitTeamNum4.get().isdigit():
        summarizeData(int(submitTeamNum4.get()))
        tableList.append(rawData)
    if submitTeamNum5.get().isdigit():
        summarizeData(int(submitTeamNum5.get()))
        tableList.append(rawData)
    if submitTeamNum6.get().isdigit():
        summarizeData(int(submitTeamNum6.get()))
        tableList.append(rawData)
    #print(tableList)

    dataTable = ttk.Treeview(tableWindow, show= "headings")
    dataTable["columns"] = ("TeamNum",  "CL1A", "CL2A",  "CL3A",  "CL4A",  "CPA",  "CNA",  "aMissed", "avgAuto", "CL1",  "CL2",  "CL3",  "CL4",  "CP",  "CN",  "tMissed" , "avgTeleop",  "defence",  "climbAcc",  "climbing",  "reefAlgae", "avgScore",)

    dataTable.column("TeamNum", width=50)
    dataTable.heading("TeamNum", text = "Team #")

    dataTable.column("CL1A", width=40)
    dataTable.heading("CL1A", text = "CL1A")
    
    dataTable.column("CL2A", width=40)
    dataTable.heading("CL2A", text = "CL2A")

    dataTable.column("CL3A", width=40)
    dataTable.heading("CL3A", text = "CL3A")

    dataTable.column("CL4A", width=40)
    dataTable.heading("CL4A", text = "CL4A")

    dataTable.column("CPA", width=40)
    dataTable.heading("CPA", text = "CPA")

    dataTable.column("CNA", width=40)
    dataTable.heading("CNA", text = "CNA")
   
    dataTable.column("aMissed", width=75)
    dataTable.heading("aMissed", text = "Auto Missed")

    dataTable.column("avgAuto", width=85)
    dataTable.heading("avgAuto", text = "Avg. Auto %")

    dataTable.column("CL1", width=40)
    dataTable.heading("CL1", text = "CL1")
    
    dataTable.column("CL2", width=40)
    dataTable.heading("CL2", text = "CL2")

    dataTable.column("CL3", width=40)
    dataTable.heading("CL3", text = "CL3")

    dataTable.column("CL4", width=40)
    dataTable.heading("CL4", text = "CL4")

    dataTable.column("CP", width=40)
    dataTable.heading("CP", text = "CP")

    dataTable.column("CN", width=40)
    dataTable.heading("CN", text = "CN")
   
    dataTable.column("tMissed", width=95)
    dataTable.heading("tMissed", text = "Teleop Missed %")

    dataTable.column("avgTeleop", width=75)
    dataTable.heading("avgTeleop", text = "Avg. Teleop")

    dataTable.column("defence", width=70)
    dataTable.heading("defence", text = "Defence %")
    
    dataTable.column("climbAcc", width=40)
    dataTable.heading("climbAcc", text = "Climb Acc.")

    dataTable.column("climbing", width=60)
    dataTable.heading("climbing", text = "climb")

    dataTable.column("avgScore", width=80)
    dataTable.heading("avgScore", text = "Avg. Score")

    dataTable.column("reefAlgae", width=40)
    dataTable.heading("reefAlgae", text = "AoR")

    for i in range(len(tableList)):
        dataTable.insert(parent='', index=i, values=tableList[i])
    dataTable.pack()

title = tk.Label(
    text= "Summarizer App",
    fg ="black"
)

openFile = tk.Button(
    root,
    text='Open Spreadsheet File',
    command=selectFile
)

fileText = tk.Label(
    text = "No File Selected"
)

enterTeams = tk.Label(
    text= "Which team would you like data for?",
    fg ="black"
)

d1 = tk.Radiobutton(
    root,
    text="Last 6 Matches",
    variable = matchAmount,
    value= "6"
)
d2 = tk.Radiobutton(
    root,
    text="All Matches",
    variable = matchAmount,
    value= "A"
)
errorOutput = tk.Label(
    text = "",
    wraplength=450
)

submitTeamNum1 = tk.Entry()
submitTeamNum2 = tk.Entry()
submitTeamNum3 = tk.Entry()
submitTeamNum4 = tk.Entry()
submitTeamNum5 = tk.Entry()
submitTeamNum6 = tk.Entry()

button = tk.Button(text = "Get Data", command = getData)
allTeamsButton = tk.Button(text = "All Teams Summary (close excel first)", command = fullSummary)

title.pack()
fileText.pack()
openFile.pack()
enterTeams.pack()
d1.pack()
d2.pack()
submitTeamNum1.pack()
submitTeamNum2.pack()
submitTeamNum3.pack()
submitTeamNum4.pack()
submitTeamNum5.pack()
submitTeamNum6.pack()
button.pack()
allTeamsButton.pack()
errorOutput.pack()

root.mainloop()